"""XLSX -> 텍스트 변환."""

from __future__ import annotations

from pathlib import Path


def convert_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl가 필요합니다. requirements-optional.txt 참고") from exc
    wb = load_workbook(path, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            row_vals = [str(cell) if cell is not None else "" for cell in row]
            if not any(val.strip() for val in row_vals):
                continue
            lines.append(" | ".join(row_vals))
        lines.append("")
    return "\n".join(lines).strip()
