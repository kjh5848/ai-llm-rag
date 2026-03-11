"""
[실습 2 - 비정형 Excel 추출]
병합된 헤더(계층형 헤더)와 문서처럼 사용된 결재란을 처리하는 스크립트입니다.
"""
import os
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_BUDGET = os.path.join(BASE_DIR, '../data/docs/finance/FIN_부서별_예산기안서.xlsx')
INPUT_FILE = os.path.join(BASE_DIR, '../../data/docs/fin/FIN_매출현황_v1.0.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, '../../parsed_data/success_results/02_FIN_매출현황_success.md')

def parse_document_excel_openpyxl(xlsx_path):
    print(f"\n📊 엑셀 문서형 추출 (Openpyxl) 시작: {os.path.basename(xlsx_path)}")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # 좌표 기반 메타데이터 추출 (Cell slicing)
    title = ws["A1"].value
    dept = ws["B4"].value
    date_val = ws["F4"].value
    
    md_output = f"# {title}\n- 기안부서: {dept}\n- 작성일자: {date_val}\n\n"
    
    # 팁: 문서 내 삽입된 이미지(차트/도형) 처리 로직 (Vision LLM 파이프라인 연계용)
    if getattr(ws, 'images', None) and len(ws.images) > 0:
        md_output += f"> 💡 [팁] 이 시트에는 {len(ws.images)}개의 이미지가 포함되어 있습니다. openpyxl로 추출 후 Vision LLM에게 전달할 수 있습니다.\n\n"
    
    return md_output

def parse_complex_header_pandas(xlsx_path):
    print(f"\n📊 엑셀 다중헤더 추출 (Pandas) 시작: {os.path.basename(xlsx_path)}")
    # Multi-Index (3개의 계층적 헤더) 로드
    df = pd.read_excel(xlsx_path, index_col=[0, 1, 2])
    
    # Markdown으로 변환
    md_table = df.to_markdown()
    return f"## 상반기 매출 현황 (다중 인덱스 표)\n\n{md_table}\n"

def process_all_excel():
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    
    result = ""
    # 1. 문서형 엑셀 (결재란 등)
    if os.path.exists(INPUT_BUDGET):
        result += parse_document_excel_openpyxl(INPUT_BUDGET)
        # 본문 테이블은 Pandas로 iloc를 써서 자를 수 있음
        df_table = pd.read_excel(INPUT_BUDGET, header=7) # 8번째 줄을 헤더로
        result += "### 예산 상세표\n" + df_table.to_markdown(index=False) + "\n\n"
        
    # 2. 다중 헤더 엑셀
    if os.path.exists(INPUT_SALES):
        result += parse_complex_header_pandas(INPUT_SALES)
        
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(result)
        
    print(f"✅ 마크다운 변환 완료: {os.path.basename(OUTPUT_FILE)}")

if __name__ == "__main__":
    process_all_excel()
