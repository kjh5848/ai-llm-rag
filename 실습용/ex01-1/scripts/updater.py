import os
import re

file_path = "0_generate_mock_docs.py"

with open(file_path, "r", encoding="utf-8") as f:
    content = f.read()

# 1. Add playwright import and html_to_pdf function at the top
replacement_top = """from PIL import Image as PilImage, ImageDraw, ImageFont

def html_to_pdf_playwright(html_path, pdf_path):
    import time
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("Playwright is not installed. Skipping PDF generation.")
        return
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            url = f"file://{os.path.abspath(html_path)}"
            page.goto(url)
            time.sleep(1)
            page.pdf(path=pdf_path, format="A4", print_background=True, margin={"top": "0cm", "bottom": "0cm", "left": "0cm", "right": "0cm"})
            browser.close()
    except Exception as e:
        print(f"Playwright PDF Error: {e}")
"""
content = re.sub(r"from PIL import Image as PilImage, ImageDraw, ImageFont", replacement_top, content)

# 2. Rewrite create_hr_rules_pdf
hr_rules = """def create_hr_rules_pdf():
    html_path = os.path.join(BASE_DIR, 'hr', 'HR_취업규칙_v1.0.html')
    pdf_path = os.path.join(BASE_DIR, 'hr', 'HR_취업규칙_v1.0.pdf')
    
    # Read HTML template from ex01
    template_path = os.path.abspath(os.path.join(BASE_DIR, '../../../../ex01/docs/html/HR_사내규정_다단.html'))
    with open(template_path, 'r', encoding='utf-8') as tf:
        html_content = tf.read()
    
    # Modify it to match current document name
    html_content = html_content.replace('사내 인사 규정 (발췌)', '취업규칙 (다단 편집형)')
    html_content = html_content.replace('Metacoding Inc.', '이노베이션 주식회사')
    
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    html_to_pdf_playwright(html_path, pdf_path)
    print(f"생성 완료: {pdf_path}")
"""
content = re.sub(r"def create_hr_rules_pdf\(\):.*?(?=def create_fin_sales_excel)", hr_rules, content, flags=re.DOTALL)

# 3. Enhance create_fin_sales_excel
fin_excel = """def create_fin_sales_excel():
    path = os.path.join(BASE_DIR, 'finance', 'FIN_2025_상반기_매출현황.xlsx')
    
    import pandas as pd
    import numpy as np
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

    arrays = [
        ['Global'] * 12,
        ['아태지역(APAC)'] * 4 + ['미주(NA)'] * 4 + ['유럽(EMEA)'] * 4,
        ['한국(서울)', '한국(부산)', '일본(도쿄)', '대만(타이베이)', '미국(뉴욕)', '미국(텍사스)', '캐나다(밴쿠버)', '멕시코(시티)', '영국(런던)', '프랑스(파리)', '독일(베를린)', '이탈리아(로마)']
    ]
    index = pd.MultiIndex.from_tuples(list(zip(*arrays)), names=['Business Unit', 'Region', 'Branch'])
    
    columns_arrays = [['1분기(Q1)'] * 3 + ['2분기(Q2)'] * 3, ['1월', '2월', '3월', '4월', '5월', '6월']]
    columns = pd.MultiIndex.from_tuples(list(zip(*columns_arrays)), names=['Quarter', 'Month'])

    df = pd.DataFrame(np.random.randint(50000, 999999, (12, 6)), index=index, columns=columns)
    
    writer = pd.ExcelWriter(path, engine='openpyxl')
    df.to_excel(writer, sheet_name='Sales_Data')
    worksheet = writer.sheets['Sales_Data']
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF")
    idx_fill = PatternFill("solid", fgColor="DCE6F1")
    idx_font = Font(bold=True, color="1F497D")
    
    for row in worksheet.iter_rows():
        for cell in row:
            if type(cell).__name__ != 'MergedCell':
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row <= 3 and cell.column > 3:
                    cell.fill = header_fill
                    cell.font = header_font
                elif cell.row > 3 and cell.column <= 3:
                    cell.fill = idx_fill
                    cell.font = idx_font
                elif cell.row > 3 and cell.column > 3:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    idx_names = ["Business Unit", "Region", "Branch"]
    for i, name in enumerate(idx_names, 1):
        cell = worksheet.cell(row=3, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col, width in zip('ABCDEFGHI', [15, 20, 20, 15, 15, 15, 15, 15, 15]):
        worksheet.column_dimensions[col].width = width

    writer.close()
    print(f"생성 완료: {path}")
"""
content = re.sub(r"def create_fin_sales_excel\(\):.*?(?=def create_fin_budget_excel)", fin_excel, content, flags=re.DOTALL)


# 4. Add create_hr_security_scan_pdf before create_ops_image definition if it was deleted, or update main block
new_main = """def create_hr_security_scan_pdf():
    html_path = os.path.join(BASE_DIR, 'hr', 'HR_정보보안서약서.html')
    pdf_path = os.path.join(BASE_DIR, 'hr', 'HR_정보보안서약서.pdf')
    
    # Read HTML template from ex01
    template_path = os.path.abspath(os.path.join(BASE_DIR, '../../../../ex01/docs/html/SEC_보안규정_스캔.html'))
    with open(template_path, 'r', encoding='utf-8') as tf:
        html_content = tf.read()
    
    html_content = html_content.replace('보안 규정 지침서', '정보보안 서약서 (스캔본)')
    html_content = html_content.replace('2026-SEC-001', '2026-HR-SEC-002')
    
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    html_to_pdf_playwright(html_path, pdf_path)
    print(f"생성 완료: {pdf_path}")

if __name__ == "__main__":
    print("--- 실무형 모의 문서 생성기 시작 ---")
    create_hr_rules_pdf()
    create_hr_security_scan_pdf()
    create_fin_sales_excel()
    create_fin_budget_excel()
    create_sec_rules_docx()
    create_ops_strategy_pdf()
    create_ops_image()
    print("--- 생성 완료 ---")
"""
content = re.sub(r"if __name__ == .__main__.:.*", new_main, content, flags=re.DOTALL)

with open(file_path, "w", encoding="utf-8") as f:
    f.write(content)
print("File updated successfully.")
