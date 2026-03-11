import pandas as pd
from openpyxl import load_workbook
import os

# 파일 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, "../../../../data/docs/ex_excel/", "05_case1_색상분류.xlsx")
output_dir = os.path.join(current_dir, "../../../../data/processed/")
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "05_case1_색상분류_성공.json")

print(f"Reading {input_file}...")

# 1. openpyxl로 워크북 열기 (data_only=True로 수식 값 읽기 권장, 스타일은 별도 접근)
wb = load_workbook(input_file, data_only=True)
ws = wb.active

# 2. 데이터 추출
data = []
s_row = 2 # 헤더 제외, 2행부터 시작

# 색상 코드 매핑 (openpyxl은 ARGB 형태로 반환됨)
COLOR_MAP = {
    "FFFF0000": "Critical (위험)",
    "FFFFFF00": "Warning (지연)",
    "00000000": "Normal",
    None: "Normal"
}

# 3. 행 순회
for row in ws.iter_rows(min_row=s_row, values_only=False):
    # values_only=False여야 cell 객체(스타일)에 접근 가능
    
    # 데이터 값 추출 (cell.value)
    row_data = {
        "id": row[0].value,
        "project_name": row[1].value,
        "manager": row[2].value,
        "deadline": str(row[3].value), # 날짜 객체 문자열 변환
        "status_text": row[4].value
    }
    
    # 4. 스타일(색상) 추출
    status_cell = row[4] # E열 (상태)
    fill_color = status_cell.fill.start_color.index # 색상 코드 (ARGB)
    
    # 색상에 따른 의미 매핑
    # 00000000 or Theme colors might be tricky, but basic fills work well
    visual_status = COLOR_MAP.get(fill_color, "Unknown")
    
    # 만약 매핑되지 않은 색상이면 코드를 그대로 저장 (추후 분석용)
    if visual_status == "Unknown":
        visual_status = f"Color({fill_color})"
        
    row_data["visual_flag"] = visual_status
    
    data.append(row_data)

# 5. DataFrame 변환 및 Markdown 저장
df = pd.DataFrame(data)

# Save to Markdown for RAG context
md_output = os.path.join(output_dir, "05_case1_색상분류_성공.md")
with open(md_output, "w", encoding="utf-8") as f:
            f.write("# [Document Success] 05_case1_색상분류 - Color to Text Conversion\n\n")
            f.write("### ✅ RAG 최적화 분석\n")
            f.write("- **전략**: 시각적 속성(Color)을 텍스트(Text)로 변환하여 검색 가능한 메타데이터로 만듦.\n")
            f.write("- **효과**: 빨간색( #FF0000)이 'Critical'이라는 명시적 상태값으로 변환되어, 의미 기반 검색이 가능해짐.\n\n")
            f.write("---\n")
            f.write("### 📄 정제된 텍스트 결과 (Sample)\n\n")
            f.write(df.head(10).to_markdown(index=False))
print(f"  -> Saved to {os.path.basename(md_output)}")
