import pandas as pd
from openpyxl import load_workbook
import os

# 파일 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, "../../../../data/docs/ex_excel/", "07_case1_수식계산.xlsx")
output_dir = os.path.join(current_dir, "../../../../data/processed/")
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "07_case1_수식계산_실패.md")

print(f"Reading {input_file}...")

# 1. 수식 그대로 읽기 (data_only=False)
# openpyxl 기본값은 False(수식 읽기)임
wb = load_workbook(input_file, data_only=False)
ws = wb.active

data = []
header = [cell.value for cell in ws[1]]
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(list(row))

df = pd.DataFrame(data, columns=header)

print("--- [Fail Load] (Formulas as Strings) ---")
print(df[["품목", "공급가액(수식)", "총액(수식)"]].head())

# 2. 결과 저장
with open(output_file, "w", encoding="utf-8") as f:
    f.write("# [Document Fail] 07_case1_수식계산 - Formula Strings Only\n\n")
    f.write("### ❌ RAG 실패 원인 분석\n")
    f.write("- **증상**: 계산된 값(예: 10000) 대신 `=B2*C2`와 같은 수식 문자열이 반환됨.\n")
    f.write("- **원인**: `openpyxl` 로드 시 `data_only=True` 옵션을 사용하지 않아, 계산 결과 값이 아닌 수식(Formula) 자체를 읽어옴.\n\n")
    f.write("---\n")
    f.write("### 📄 정제된 텍스트 결과 (Sample)\n")
    f.write(df.to_markdown(index=False))

print(f"  -> Saved failure case to {os.path.basename(output_file)}")
