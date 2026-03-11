import pandas as pd
from openpyxl import load_workbook
import os

# 파일 경로 설정
current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, "../../../../data/docs/ex_excel/", "08_case1_숨겨진행.xlsx")
output_dir = os.path.join(current_dir, "../../../../data/processed/")
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, "08_case1_숨겨진행_성공.json")

print(f"Reading {input_file}...")

# 1. 일반적인 읽기 (Pandas) -> 숨겨진 데이터도 포함됨
df_standard = pd.read_excel(input_file)
print("--- [Standard Load] (Includes Hidden Rows) ---")
print(df_standard) 
# 예상: P099, P100 ("숨김처리됨") 데이터가 보임 -> RAG 오염 원인

# 2. 숨겨진 행 제외하고 읽기 (openpyxl)
wb = load_workbook(input_file, data_only=True)
ws = wb.active

data = []
header = [cell.value for cell in ws[1]] # 첫 번째 행은 헤더

# 2행부터 순회
for row in ws.iter_rows(min_row=2):
    # 행 숨김 여부 확인
    row_dim = ws.row_dimensions[row[0].row]
    if row_dim.hidden:
        continue # 숨겨진 행 스킵
    
    # 데이터 추출
    row_values = [cell.value for cell in row]
    data.append(row_values)

# DataFrame 생성
df_filtered = pd.DataFrame(data, columns=header)
print("\n--- [Filtered Load] (Excludes Hidden Rows) ---")
print(df_filtered)

# Save to Markdown
md_output = os.path.join(output_dir, "08_case1_숨겨진행_성공.md")
with open(md_output, "w", encoding="utf-8") as f:
    f.write("# [Document Success] 08_case1_숨겨진행 - Filtering Hidden Rows\n\n")
    f.write("### ✅ RAG 최적화 분석\n")
    f.write("- **전략**: `openpyxl`을 사용하여 숨겨진 행(Hidden Rows)을 감지하고 제외(Filter Out) 처리.\n")
    f.write("- **효과**: 사용자에게 노출되지 않는(폐기된) 데이터가 검색 결과에 포함되는 Hallucination 방지.\n\n")
    f.write("---\n")
    f.write("### 📄 정제된 텍스트 결과 (Sample)\n\n")
    f.write(df_filtered.to_markdown(index=False))

print(f"  -> Saved to {os.path.basename(md_output)}")
